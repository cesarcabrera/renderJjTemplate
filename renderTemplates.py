#!/usr/bin/env python3

# Copyright (C) 2025 César Cabrera / cesarcabrera.info@gmail.com
# This is free software: you can redistribute it and/or modify it
# under the terms of the 3-clause BSD License.
# renderjj2Templates is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY of any kind whatsoever.

import jinja2
import pandas as pd
import openpyxl
import json
import re
import argparse
from datetime import datetime
import os

LATEST_VERSION = '0.1'
DEB_LEVEL = 1


def readTemplates(path='.', filename='configTemplates.xlsx', sheet_name='Sheet1'):
    """ Reads an Excel file and stores it into strings to be used as jj2 templates
        Detects {{}} patterns and appends it to variables
    """

    df = pd.read_excel(path + "/" + filename, sheet_name=sheet_name)
    template = ''
    variables = []
    for row in df.iterrows():
        linea = str(row[1].values[0]) + '\n'
        var = re.findall(r"\{\{([\w-]+)\}\}", linea)
        if len(var) > 1:
            variables += var
        template += linea
    return template, variables


def readVariables(path='.', filename='configVariables.xlsx', sheet_name='Sheet1'):
    """ Reads an Excel file and stores it into a dictionary to be used as jj2 data
        Detect values starting with { and reads it as json data (not secure)
    """

    df = pd.read_excel(path + "/" + filename, sheet_name=sheet_name)
    data = {}
    for row in df.iterrows():
        name, value = row[1][0], row[1].values[1]
        linea = str(value)
        data[name] = ''
        if re.match(r'\{', linea):
            print(f"DEB: i = {row[0]}, name = {name} (Reading JSON)") if DEB_LEVEL >= 1 else ""
            content = json.loads(linea)
        else:
            print(f"DEB: i = {row[0]}, name = {name} (Reading Text)") if DEB_LEVEL >= 1 else ""
            content = linea
        data[name] = content
        print(f"DEB: {name} => {data[name]}") if DEB_LEVEL >= 2 else ""
    return data


if __name__ == '__main__':
    start = datetime.now()
    date_prefix = start.strftime('%Y_%m_%d_%H_%M')

    parser = argparse.ArgumentParser()
    parser.add_argument('-v', '--version', help='Version', action='version',
                        version="cesarcabrera.info 2025 - Version: " + LATEST_VERSION)
    parser.add_argument('-f', '--folder', help='Config+parameters+output folder. Default: .', default='.')
    parser.add_argument('-c', '--config', help='Config Filename. Def. configTemplates.xlsx',
                        default='configTemplates.xlsx')
    parser.add_argument('-p', '--parameters', help='Config variables (parameters) file. Def. configVariables.xlsx',
                        default='configVariables.xlsx')
    parser.add_argument('-o', '--output', help='Filenames prefix. Def. OUT_', default='OUT_')
    parser.add_argument('-u', '--unique', help='Use only the 1st sheet from config file for all the variable sheets '
                                               '(one template for all sites)',
                        action='store_true')
    parser.add_argument('-g', '--section-groups', help='Select the groups to show (Sheet name from variables file) '
                                                       'Do not use "-" nor spaces or special chars, please',
                        default='Sheet1')
    parser.add_argument('-s', '--show-vars', help='(Optional)Show detected variables', action='store_true')
    parser.add_argument('-t', '--show-template', help='(Optional)Show templates', action='store_true')
    parser.add_argument('-i', '--interactive', help='(Optional)Wait for after showing one rendered template',
                        action='store_true')
    parser.add_argument('-k', '--cat', help='(Optional)Write all on one single file',
                        action='store_true')
    parser.add_argument('-d', '--debug', help='Debug Level (0=Nothing, > 1 more detail)', type=int)
    parser.add_argument('-x', help='Consolidate ouput on 1 file', action='store_true')

    args = parser.parse_args()
    DEB_LEVEL = args.debug if args.debug else 0

    templateFolder, systemSeparator = args.folder, '/'
    configFile, variablesFile = args.config, args.parameters

    if templateFolder != '.' and not os.path.isdir(templateFolder):
        print(f"INFO: Creating {templateFolder}")
        os.mkdir(templateFolder)

    wb = openpyxl.load_workbook(templateFolder + systemSeparator + variablesFile)
    sectionGroups = wb.sheetnames
    print(f"INFO: detected {len(sectionGroups)} sectionGroups")
    print(f"INFO: {sectionGroups}")
    env = jinja2.Environment()
    uniqueRead, uniqueTemplate = True, ''
    if args.unique:
        wb = openpyxl.load_workbook(templateFolder + systemSeparator + configFile)
        uniqueTemplate = wb.sheetnames[0]
        uniqueRead = False
    temp, var = "", {}
    consolidated = ""
    for g in sectionGroups:
        if args.section_groups != "Sheet1":
            if g != args.section_groups:
                continue
        print(f"INFO: >>> {g}")
        if args.unique:
            temp, var = readTemplates(path=templateFolder, filename=configFile, sheet_name=uniqueTemplate)
            uniqueRead = True
        else:
            temp, var = readTemplates(path=templateFolder, filename=configFile, sheet_name=g)
        data = readVariables(path=templateFolder, filename=variablesFile, sheet_name=g)
        if args.show_vars:
            print(f"INFO: Detected variables = {data}")
        template = env.from_string(str(temp))
        if args.show_template:
            print(f"INFO: >>> {g} TEMPLATE")
            print(str(temp))
            print(f"INFO: <<< {g} TEMPLATE")
        config = template.render(data)
        outputFilename = templateFolder + systemSeparator + args.output + date_prefix + g + '.txt'
        if os.path.isfile(outputFilename):
            input(f"WARN: {outputFilename} already exists and will be overwitten. Ctrl-C to stop processing")
        with open(outputFilename, 'w+') as f:
            print(f"INFO:  Writing to {outputFilename} {len(config)} lines!")
            f.write(config)
            consolidated += "\n" + config
        print(config) if DEB_LEVEL >= 2 else ""
        print(f"INFO: <<< {g}")
        if args.interactive:
            input("Press ENTER to continue...")
    if args.x:
        outputFilename = templateFolder + systemSeparator + args.output + date_prefix + ".txt"
        print(f"INFO: Consolidating output in {outputFilename}")
        with open(outputFilename, 'w') as f:
            f.write(consolidated)
    last = datetime.now() - start
    print(f"INFO: Execution took {last.microseconds} microseconds")
